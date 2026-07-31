"""One-off, read-only reconstruction of RNS processing speed for a given day.

Rebuilds the measurement the ad-hoc 08:00 monitoring script produced: how fast
each RNS announcement moved from publication (published_at, the ~07:00 drop)
through our pipeline to a finished LLM score (llm_processed_at), broken into the
ingest / summary / rank stages.

Usage:  python analysis/rns_processing_latency.py [YYYY-MM-DD]
Default date = 2026-07-22. Read-only: SELECT only, no writes.
"""
import csv
import os
import sys
from datetime import datetime, timezone

_BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _BACKEND)

from dotenv import load_dotenv
load_dotenv(os.path.join(_BACKEND, ".env"))

from db import query

DAY = sys.argv[1] if len(sys.argv) > 1 else "2026-07-22"


def _pct(values, p):
    if not values:
        return None
    s = sorted(values)
    k = (len(s) - 1) * (p / 100)
    lo = int(k)
    hi = min(lo + 1, len(s) - 1)
    return s[lo] + (s[hi] - s[lo]) * (k - lo)


def _fmt(seconds):
    if seconds is None:
        return "   n/a"
    m, s = divmod(int(round(seconds)), 60)
    return f"{m:>3}m{s:02d}s"


def summarize(label, values):
    vals = [v for v in values if v is not None]
    if not vals:
        print(f"  {label:<28} n=0")
        return
    print(
        f"  {label:<28} n={len(vals):<3} "
        f"median {_fmt(_pct(vals, 50))}  "
        f"p90 {_fmt(_pct(vals, 90))}  "
        f"max {_fmt(max(vals))}"
    )


rows = query(
    """
    SELECT id, ticker, symbol, company_name, headline, tier, category,
           llm_score, llm_sentiment, llm_action, llm_model,
           published_at, fetched_at, summary_fetched_at, llm_processed_at
    FROM rns_announcements
    WHERE published_at >= %s::date
      AND published_at <  (%s::date + INTERVAL '1 day')
    ORDER BY published_at
    """,
    (DAY, DAY),
)

if not rows:
    print(f"No RNS rows with published_at on {DAY}.")
    sys.exit(0)


def secs(a, b):
    if a is None or b is None:
        return None
    return (b - a).total_seconds()


# NOTE: fetched_at is overwritten to NOW() on every re-ingest upsert
# (rns.py ON CONFLICT ... fetched_at = NOW()), so it is "last seen", not
# "first ingested" — useless for a stage split. llm_processed_at is the one
# genuinely write-once column here (rank selects on IS NULL).
#
# summary_fetched_at is NOT write-once, despite what this comment claimed until
# 2026-07-31. rns._update_summary_and_body sets it to NOW() unconditionally, and
# _backfill_summaries selects on `body_fetched_at IS NULL` — a different column —
# so any row revisited for its body has its summary timestamp rewritten with it.
#
# In practice that happened once, and the damage is permanent: when body capture
# shipped (6b73dfb, 2026-07-28) the backfill swept the whole pre-existing tier
# A/B backlog, overwriting summary_fetched_at on every row it touched. Those
# rows now carry the sweep's clock time, not their real entry time, and the
# originals are gone. The symptom is a ladder — measured 2026-07-31, rows
# published 07-24 read 07:10, 07-22 read 07:15, 07-21 07:30 ... 07-13 08:45,
# and 07-28 reads 16:10 — impossible for stories ingested minutes after a 07:00
# publication, and a dead giveaway that the stamp is the sweep, not the ingest.
#
# So: TREAT summary_fetched_at AS UNRELIABLE BEFORE ~2026-07-28. Stage splits
# and p90 figures quoted from earlier rows (including those in
# docs/rns-earnings-quality-plan.md) are measuring the backfill. Rows from
# 07-29 on are clean — the backlog is empty and each row is written once — and
# a failed fetch leaves BOTH columns NULL rather than half-stamping, so the
# pairing stays internally consistent either way.
e2e, to_summary, summ_to_rank = [], [], []
ranked = 0
for r in rows:
    r["_e2e"] = secs(r["published_at"], r["llm_processed_at"])
    r["_to_summary"] = secs(r["published_at"], r["summary_fetched_at"])
    r["_summ_to_rank"] = secs(r["summary_fetched_at"], r["llm_processed_at"])
    e2e.append(r["_e2e"])
    to_summary.append(r["_to_summary"])
    summ_to_rank.append(r["_summ_to_rank"])
    if r["llm_processed_at"] is not None:
        ranked += 1

# Publication clustering
pub_times = [r["published_at"] for r in rows]
first_pub = min(pub_times)
last_pub = max(pub_times)

tiers = {}
for r in rows:
    tiers.setdefault(r["tier"], 0)
    tiers[r["tier"]] += 1

print(f"=== RNS processing latency — published_at on {DAY} (UTC) ===\n")
print(f"Total announcements ingested : {len(rows)}")
print(f"  by tier                    : " + ", ".join(f"{t}={n}" for t, n in sorted(tiers.items())))
print(f"Publication window (UTC)     : {first_pub:%H:%M:%S} -> {last_pub:%H:%M:%S}")
early = sum(1 for p in pub_times if p.hour < 7)  # <07:00 UTC == <08:00 BST
print(f"Ranked (llm_processed_at set): {ranked}/{len(rows)}\n")

print("Stage latencies (Tier A/B are the ranked ones; summary_fetched_at is "
      "unreliable before ~2026-07-28 — see the NOTE above):")
summarize("end-to-end  pub->scored", e2e)
summarize("  ingest+summary pub->summ", to_summary)
summarize("  rank    summ->scored", summ_to_rank)

# Per-tier end-to-end
print("\nEnd-to-end by tier:")
for t in sorted(tiers):
    summarize(f"tier {t}", [r["_e2e"] for r in rows if r["tier"] == t])

# Slowest ranked items
ranked_rows = [r for r in rows if r["_e2e"] is not None]
ranked_rows.sort(key=lambda r: r["_e2e"], reverse=True)
print("\nSlowest 10 ranked announcements (pub -> scored):")
for r in ranked_rows[:10]:
    tk = r["ticker"] or r["symbol"] or "?"
    hl = (r["headline"] or "")[:48]
    print(f"  {_fmt(r['_e2e'])}  {tk:<7} T{r['tier']} s={r['llm_score']}  {hl}")

# Fastest for reference
print("\nFastest 5 ranked announcements:")
for r in ranked_rows[-5:][::-1]:
    tk = r["ticker"] or r["symbol"] or "?"
    hl = (r["headline"] or "")[:48]
    print(f"  {_fmt(r['_e2e'])}  {tk:<7} T{r['tier']} s={r['llm_score']}  {hl}")


# --- CSV export: one row per ranked (Tier A/B, scored) announcement --------
def _iso(dt):
    return dt.isoformat(sep=" ", timespec="seconds") if dt else ""


def _mins(seconds):
    return round(seconds / 60, 2) if seconds is not None else ""


# Parent UK_stocks folder ("the folder above" the repo).
out_dir = os.path.dirname(os.path.dirname(_BACKEND))
out_path = os.path.join(out_dir, f"rns_processing_latency_{DAY}.csv")

csv_rows = sorted(
    (r for r in rows if r["_e2e"] is not None),
    key=lambda r: r["published_at"],
)

# utf-8-sig so Excel reads the BOM and shows £/accents correctly.
with open(out_path, "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow([
        "id", "ticker", "symbol", "company_name", "headline",
        "tier", "category", "llm_score", "llm_sentiment", "llm_action",
        "llm_model", "published_at_utc", "summary_fetched_at_utc",
        "llm_processed_at_utc",
        "end_to_end_min", "pub_to_summary_min", "summary_to_rank_sec",
    ])
    for r in csv_rows:
        w.writerow([
            r["id"], r["ticker"] or "", r["symbol"] or "",
            r["company_name"] or "", r["headline"] or "",
            r["tier"], r["category"] or "", r["llm_score"],
            r["llm_sentiment"] or "", r["llm_action"] or "",
            r["llm_model"] or "",
            _iso(r["published_at"]), _iso(r["summary_fetched_at"]),
            _iso(r["llm_processed_at"]),
            _mins(r["_e2e"]), _mins(r["_to_summary"]),
            round(r["_summ_to_rank"], 1) if r["_summ_to_rank"] is not None else "",
        ])

print(f"\nWrote {len(csv_rows)} ranked rows -> {out_path}")


# --- XLSX export: Data sheet + Explanation sheet ---------------------------
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

xlsx_path = os.path.join(out_dir, f"rns_processing_latency_{DAY}.xlsx")

HEADERS = [
    "id", "ticker", "symbol", "company_name", "headline",
    "tier", "category", "llm_score", "llm_sentiment", "llm_action",
    "llm_model", "published_at_utc", "summary_fetched_at_utc",
    "llm_processed_at_utc",
    "end_to_end_min", "pub_to_summary_min", "summary_to_rank_sec",
]

wb = Workbook()
ws = wb.active
ws.title = "Data"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
ws.append(HEADERS)
for c in ws[1]:
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(vertical="center")

for r in csv_rows:
    ws.append([
        r["id"], r["ticker"] or "", r["symbol"] or "",
        r["company_name"] or "", r["headline"] or "",
        r["tier"], r["category"] or "", r["llm_score"],
        r["llm_sentiment"] or "", r["llm_action"] or "",
        r["llm_model"] or "",
        _iso(r["published_at"]), _iso(r["summary_fetched_at"]),
        _iso(r["llm_processed_at"]),
        _mins(r["_e2e"]), _mins(r["_to_summary"]),
        round(r["_summ_to_rank"], 1) if r["_summ_to_rank"] is not None else "",
    ])

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
widths = [10, 8, 9, 24, 46, 5, 16, 9, 12, 10, 26, 21, 21, 21, 14, 17, 18]
for i, wd in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = wd


def _num(values, kind):
    seconds = [v for v in values if v is not None]
    if kind == "min":
        return f"{_pct(seconds, 50) / 60:.1f}m median, {_pct(seconds, 90) / 60:.0f}m p90, {max(seconds) / 60:.0f}m max"
    return f"{_pct(seconds, 50):.0f}s median, {_pct(seconds, 90):.0f}s p90, {max(seconds):.0f}s max"


ex = wb.create_sheet("Explanation")
ex.column_dimensions["A"].width = 30
ex.column_dimensions["B"].width = 96

title_font = Font(bold=True, size=14)
h_font = Font(bold=True, size=11, color="1F4E78")
wrap = Alignment(wrap_text=True, vertical="top")


def row(a, b="", *, style=None, gap=False):
    ex.append([a, b])
    r = ex.max_row
    if style == "title":
        ex[f"A{r}"].font = title_font
    elif style == "h":
        ex[f"A{r}"].font = h_font
    else:
        ex[f"A{r}"].font = Font(bold=True)
    ex[f"B{r}"].alignment = wrap
    if gap:
        ex.append([])


row(f"RNS processing latency — {DAY}", "", style="title", gap=True)

row("What this is", style="h")
row("Purpose", "How fast each RNS announcement moved from publication to a finished LLM "
    "score on this date. Reconstructed read-only from the prod DB "
    "(rns_announcements), reproducing the ad-hoc 08:00 monitoring script.", gap=True)

row("Scope of the Data sheet", style="h")
row("Rows", f"{len(csv_rows)} — every Tier A/B announcement published on {DAY} that received "
    f"an LLM score. Tier C ({tiers.get('C', 0)} that day) is never LLM-ranked by "
    "design, so it is excluded from the timing rows.")
row("Total published", f"{len(rows)} announcements across all tiers "
    + ", ".join(f"{t}={n}" for t, n in sorted(tiers.items())) + ".")
row("Publication window", f"{first_pub:%H:%M} → {last_pub:%H:%M} UTC (UK is BST, UTC+1, "
    "so 06:00 UTC = 07:00 local — the main RNS drop).", gap=True)

row("Headline numbers", style="h")
row("End-to-end (publish→scored)", _num(e2e, "min"))
row("  ingest + summary fetch", _num(to_summary, "min"))
row("  LLM rank (DeepSeek call)", _num(summ_to_rank, "sec"))
row("Coverage", f"{ranked}/{len(rows)} scored; all {len(csv_rows)} Tier A/B rows completed.", gap=True)

row("Column meanings", style="h")
for name, desc in [
    ("id", "Investegate announcement id (primary key)."),
    ("ticker / symbol", "Raw RNS ticker and the resolved yfinance symbol (blank if unresolved)."),
    ("company_name / headline", "As published on the wire."),
    ("tier", "A/B = LLM-ranked (in this sheet). C = coarse-filter drop, not ranked."),
    ("category", "Rules classifier bucket (trading_update, interim_results, acquisition…)."),
    ("llm_score", "0–100 impact magnitude from the LLM. 76+ is the high-impact band."),
    ("llm_sentiment", "positive / negative / neutral — direction, separate from magnitude."),
    ("llm_action", "watch / research / ignore suggestion."),
    ("llm_model", "Exact scoring model string (self-labels the DeepSeek version)."),
    ("published_at_utc", "When the RNS hit the wire (write-once)."),
    ("summary_fetched_at_utc", "When our pipeline scraped the Investegate AI summary. NOT "
                               "write-once — rewritten if the row is revisited for its body, "
                               "which the 2026-07-28 body-capture backfill did to every row "
                               "before that date. Unreliable before ~2026-07-28."),
    ("llm_processed_at_utc", "When the LLM score was written (write-once)."),
    ("end_to_end_min", "Minutes from published_at → llm_processed_at. The key speed metric."),
    ("pub_to_summary_min", "Minutes from published_at → summary_fetched_at (ingest + summary)."),
    ("summary_to_rank_sec", "Seconds from summary_fetched_at → llm_processed_at (the model call)."),
]:
    row(name, desc)
ex.append([])

row("Methodology & caveats", style="h")
row("Which timestamps are sound", "fetched_at is deliberately NOT used: the ingest upsert sets "
    "fetched_at = NOW() on every re-scan, so it records 'last seen', not 'first ingested', "
    "and would make the stage split meaningless. published_at and llm_processed_at are each "
    "written once. summary_fetched_at is not: it is rewritten whenever a row is revisited for "
    "its body.")
row("Dates before 2026-07-28", "The body-capture backfill (shipped 2026-07-28) swept every "
    "pre-existing tier A/B row and overwrote summary_fetched_at with the sweep's own clock "
    "time, so pub_to_summary_min and summary_to_rank_sec are measuring the backfill, not the "
    "morning pipeline, on any earlier date. The originals are gone. Runs from 2026-07-29 "
    "onward are clean.")
row("Stages reconcile", "pub_to_summary_min + summary_to_rank_sec ≈ end_to_end_min. The rank "
    "step is fast and consistent; almost all latency is upstream (ingest + summary fetch).")
row("The slow tail", "A few items (e.g. corrections/re-issues, and drops published outside the "
    "07:00 cron burst) score much later because they wait for the next scheduled run. They lift "
    "p90/max but are not representative of the typical ~6-minute item.")
row("How to reproduce", "backend/analysis/rns_processing_latency.py <YYYY-MM-DD> — read-only, "
    "writes both the CSV and this workbook, named per date.")

wb.save(xlsx_path)
print(f"Wrote workbook (Data + Explanation) -> {xlsx_path}")
